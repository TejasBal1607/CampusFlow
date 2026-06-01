import openpyxl
import json
import os
import re
import datetime

# 🚀 SPEED HACK: Stops openpyxl from hanging on Thapar's invisible ghost formatting
import openpyxl.worksheet._reader
openpyxl.worksheet._reader.WorksheetReader.bind_merged_cells = lambda self: None

# Add your campus subject codes here over time
SUBJECT_MAP = {
    "UMA010": "Mathematics - I",
    "UPH004": "Applied Physics",
    "UTA015": "Engineering Drawing",
    "UES101": "Electrical Engineering",
    "UHU003": "Professional Communication",
    "UCB008": "Applied Chemistry",
    "UES100": "Intro to Engineering",
    "UTA013": "Engineering Design",
    "UCS100": "Intro to Computers",
    "UES102": "Solid Mechanics",
    "UHU005": "Humanities",
    "UES104": "Computer Programming" 
}

def parse_subject_and_venue(val_str: str):
    """Deep parses the cell to find Subject, Type, and Venue"""
    parts = [p.strip() for p in re.split(r'[\n\r,]+', val_str) if p.strip()]
    if not parts: parts = val_str.split()
    if not parts: return None, None, None
    
    raw_subject = parts[0]
    c_type = "Lecture"
    
    match = re.match(r'([A-Z]{3}\d{3})[\s\(\[\-]*([LPT])?[\)\]]?', raw_subject.upper())
    if match:
        base_code = match.group(1)
        modifier = match.group(2)
        if modifier == 'P': c_type = "Lab"
        elif modifier == 'T': c_type = "Tutorial"
        friendly_name = SUBJECT_MAP.get(base_code, base_code)
    else:
        friendly_name = raw_subject.upper()
        if "(P)" in friendly_name or "LAB" in friendly_name: c_type = "Lab"
        elif "(T)" in friendly_name or "TUT" in friendly_name: c_type = "Tutorial"
        
    venue = "TBA"
    for p in parts:
        p_clean = re.sub(r'\s+', '', p.upper())
        if re.search(r'^(LT\d+|LP\d+|TAN|C\-\d+|F\-\d+|G\-\d+|E\-\d+|B\-\d+|A\-\d+|D\-\d+|ROOM\d+)', p_clean):
            venue = p
            break
            
    return friendly_name, c_type, venue

def normalize_batch(batch_str: str):
    """Converts '1B1A' -> '1B11' uniformly"""
    batch_str = re.sub(r'\s+', '', str(batch_str).upper())
    mapping = {'A': '1', 'B': '2', 'C': '3', 'D': '4', 'E': '5', 'F': '6', 'G': '7', 'H': '8', 'I': '9'}
    if not batch_str: return batch_str
    if len(batch_str) >= 4 and batch_str[-1].isalpha() and batch_str[-1] in mapping and batch_str[-2].isdigit():
        return batch_str[:-1] + mapping[batch_str[-1]]
    return batch_str

def is_valid_batch(col_name: str):
    c = str(col_name).upper().strip()
    if len(c) < 3 or len(c) > 7: return False
    if not any(char.isdigit() for char in c): return False 
    if not any(char.isalpha() for char in c): return False
    junk = ['DAY', 'HOUR', 'TIME', 'SR', 'NO', 'PRACTICAL', 'UNNAMED', 'LUNCH', 'BREAK']
    if any(j in c for j in junk): return False
    return True

def format_time(t_val):
    if t_val is None: return None
    if isinstance(t_val, datetime.time):
        hr, mn = t_val.hour, t_val.minute
        period = "PM" if hr >= 12 else "AM"
        hr12 = hr - 12 if hr > 12 else (12 if hr == 0 else hr)
        return f"{hr12}:{mn:02d} {period}"
    
    t_str = str(t_val).strip()
    m = re.search(r'(\d{1,2}):(\d{2})', t_str)
    if m:
        hr, mn = int(m.group(1)), int(m.group(2))
        period = "PM" if "PM" in t_str.upper() else "AM" if "AM" in t_str.upper() else ("PM" if hr >= 12 or (1 <= hr <= 7) else "AM")
        hr12 = hr - 12 if hr > 12 else (12 if hr == 0 else hr)
        return f"{hr12}:{mn:02d} {period}"
    return t_str

def extract_excel_timetable(excel_path: str, output_file: str = "master_timetable.json"):
    print(f"📂 Loading Excel Workbook: {excel_path}...")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    master_schedule = {}
    valid_days = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY']

    for sheet_name in wb.sheetnames:
        if "PG" in sheet_name.upper() or "DLIT" in sheet_name.upper(): continue
        print(f"\n⚙️ Extracting Sheet: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # 1. Unmerge and copy cell data (Copies lectures across all sub-batches)
        for merged_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = ws.cell(row=min_row, column=min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = top_left_value

        # 2. Convert to raw Python Grid
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
        if not grid: continue

        # 3. Locate Header Row (Where "DAY" and "HOURS" exist)
        header_row_idx, day_col_idx, time_col_idx = -1, -1, -1
        for r_idx, row in enumerate(grid):
            row_strs = [str(x).upper().strip() if x is not None else "" for x in row]
            if any("DAY" in s for s in row_strs) and any("HOUR" in s or "TIME" in s for s in row_strs):
                header_row_idx = r_idx
                for c_idx, s in enumerate(row_strs):
                    if "DAY" in s: day_col_idx = c_idx
                    elif "HOUR" in s or "TIME" in s: time_col_idx = c_idx
                break
                
        if header_row_idx == -1 or day_col_idx == -1 or time_col_idx == -1:
            print("  -> ⚠️ Skipping: Could not locate 'DAY' or 'HOURS' header.")
            continue

        # 4. Map columns to Batch IDs
        batch_map = {}
        for c_idx, val in enumerate(grid[header_row_idx]):
            if c_idx in [day_col_idx, time_col_idx] or val is None: continue
            val_str = str(val).strip()
            if is_valid_batch(val_str):
                batch_map[c_idx] = normalize_batch(val_str)

        if not batch_map: continue

        # 5. Extract Data Rows
        current_day = "Monday" # The magic tracker
        extracted_count = 0

        for r_idx in range(header_row_idx + 1, len(grid)):
            row = grid[r_idx]
            
            # --- 🟢 THE DAY FIX 🟢 ---
            # Instead of strict equality, we check if the word is INSIDE the cell!
            if day_col_idx < len(row) and row[day_col_idx] is not None:
                raw_day = str(row[day_col_idx]).upper().strip()
                for d in valid_days:
                    if d in raw_day:
                        current_day = d.capitalize()
                        break
            
            if current_day == "Saturday": continue

            # Extract Time
            if time_col_idx >= len(row) or row[time_col_idx] is None: continue
            clean_time = format_time(row[time_col_idx])
            if not clean_time: continue 
            
            # Extract Classes
            for c_idx, batch_id in batch_map.items():
                if c_idx >= len(row) or row[c_idx] is None: continue
                
                cell_val = str(row[c_idx]).strip()
                if not cell_val or cell_val.upper() in ['NAN', 'NONE', '-']: continue

                subject_name, c_type, venue = parse_subject_and_venue(cell_val)
                if not subject_name: continue
                
                if batch_id not in master_schedule:
                    master_schedule[batch_id] = [{"day": d, "classes": []} for d in valid_days]
                    extracted_count += 1

                day_data = next((d for d in master_schedule[batch_id] if d["day"] == current_day), None)
                if day_data:
                    # Prevent duplicate back-to-back hour blocks of the same lab
                    if len(day_data["classes"]) > 0 and day_data["classes"][-1]["name"] == subject_name and day_data["classes"][-1]["type"] == c_type:
                        continue 
                    day_data["classes"].append({
                        "name": subject_name, 
                        "time": clean_time, 
                        "venue": venue, 
                        "type": c_type
                    })

        print(f"  -> Successfully extracted {extracted_count} batches!")

    with open(output_file, 'w') as f:
        json.dump(master_schedule, f, indent=4)
    print(f"\n✅ BOOM! Master JSON perfectly generated: {output_file}")

if __name__ == "__main__":
    EXCEL_FILENAME = "UG1,2,3 TIME TABLE JAN TO MAY 2026.xlsx" 
    if os.path.exists(EXCEL_FILENAME): extract_excel_timetable(EXCEL_FILENAME)
    else: print("❌ Excel file not found!")